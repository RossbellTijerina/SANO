from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models.oficio import OficioModel
from models.audit import AuditModel
from middleware.auth_middleware import admin_requerido
from datetime import datetime
import io
import os

reportes_bp = Blueprint('reportes', __name__)


@reportes_bp.route('/pdf', methods=['GET'])
@jwt_required()
def exportar_pdf():
    """Exportar reportes a PDF."""
    from fpdf import FPDF
    
    # Obtener filtros
    anio = request.args.get('anio', datetime.now().year, type=int)
    filtros = {
        'anio': request.args.get('anio'),
        'fecha_desde': request.args.get('fecha_desde'),
        'fecha_hasta': request.args.get('fecha_hasta'),
        'estado': request.args.get('estado'),
        'solicitante': request.args.get('solicitante'),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    
    oficios = OficioModel.buscar(filtros) if filtros else OficioModel.buscar({'anio': str(anio)})
    
    class PDF(FPDF):
        def header(self):
            # Obtener ruta del logo
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../frontend/public/assets/logo.png')
            
            # Intentar agregar logo si existe
            if os.path.exists(logo_path):
                try:
                    self.image(logo_path, 10, 8, 25)
                except Exception:
                    pass # Ignorar si FPDF no puede leer el tipo de imagen
                    
            self.set_y(15)
            self.set_font('Arial', 'B', 16)
            self.set_text_color(21, 50, 67) # #153243 Navy Blue
            self.cell(0, 8, f'SANO - REPORTE DE OFICIOS {anio}', 0, 1, 'C')
            
            self.set_font('Arial', '', 10)
            self.set_text_color(100, 100, 100)
            self.cell(0, 6, 'Gobierno del Estado de Quintana Roo', 0, 1, 'C')
            self.cell(0, 6, 'Registro Publico de la Propiedad y del Comercio - Delegacion Cancun', 0, 1, 'C')
            self.cell(0, 6, f'Fecha de Emision: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'C')
            self.ln(12)
            
            # Table Header
            self.set_fill_color(21, 50, 67) # #153243 Navy Blue
            self.set_text_color(255, 255, 255)
            self.set_draw_color(21, 50, 67)
            self.set_font('Arial', 'B', 9)
            self.cell(15, 12, 'Folio', 1, 0, 'C', 1)
            self.cell(15, 12, 'Ano', 1, 0, 'C', 1)
            self.cell(50, 12, 'Expediente', 1, 0, 'C', 1)
            self.cell(60, 12, 'Solicitante', 1, 0, 'C', 1)
            self.cell(85, 12, 'Asunto', 1, 0, 'C', 1)
            self.cell(25, 12, 'Fecha', 1, 0, 'C', 1)
            self.cell(25, 12, 'Estado', 1, 1, 'C', 1)
            self.set_text_color(50, 50, 50)

    pdf = PDF('L', 'mm', 'A4')
    pdf.add_page()
    pdf.set_font('Arial', '', 8)
    
    fill = False
    for o in oficios:
        if fill:
            pdf.set_fill_color(248, 250, 252) # light gray for alternate rows
        else:
            pdf.set_fill_color(255, 255, 255)
            
        solicitante = o['solicitante'][:30] if o['solicitante'] else ''
        asunto = o['asunto'][:45] if o['asunto'] else ''
        
        def safe_text(txt):
            return str(txt).encode('latin-1', 'replace').decode('latin-1') if txt else ''
            
        pdf.cell(15, 8, safe_text(o['folio']), 1, 0, 'C', fill)
        pdf.cell(15, 8, safe_text(o['anio']), 1, 0, 'C', fill)
        pdf.cell(50, 8, safe_text(o['expediente']), 1, 0, 'C', fill)
        pdf.cell(60, 8, safe_text(solicitante), 1, 0, 'L', fill)
        pdf.cell(85, 8, safe_text(asunto), 1, 0, 'L', fill)
        pdf.cell(25, 8, safe_text(o['fecha']), 1, 0, 'C', fill)
        pdf.cell(25, 8, safe_text(o['estado']), 1, 1, 'C', fill)
        fill = not fill
        
    pdf.ln(5)
    pdf.set_font('Arial', 'I', 9)
    pdf.cell(0, 10, f'Total de registros: {len(oficios)}', 0, 1, 'L')
    
    # Save to buffer
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    # Registrar en auditoria
    user_id = int(get_jwt_identity())
    AuditModel.registrar(user_id, 'EXPORTAR_PDF', 'Reportes', None, f"Reporte PDF generado - Ano {anio}")
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_oficios_{anio}_{datetime.now().strftime("%Y%m%d")}.pdf',
        mimetype='application/pdf'
    )


@reportes_bp.route('/excel', methods=['GET'])
@jwt_required()
def exportar_excel():
    """Exportar reporte a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    anio = request.args.get('anio', datetime.now().year, type=int)
    filtros = {
        'anio': request.args.get('anio'),
        'fecha_desde': request.args.get('fecha_desde'),
        'fecha_hasta': request.args.get('fecha_hasta'),
        'estado': request.args.get('estado'),
        'solicitante': request.args.get('solicitante'),
    }
    filtros = {k: v for k, v in filtros.items() if v}
    
    oficios = OficioModel.buscar(filtros) if filtros else OficioModel.buscar({'anio': str(anio)})
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Oficios {anio}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="153243", end_color="153243", fill_type="solid")
    title_font = Font(bold=True, size=16, color="153243")
    subtitle_font = Font(italic=True, size=11, color="555555")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'), 
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), 
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'GOBIERNO DEL ESTADO DE QUINTANA ROO'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'SANO - Reporte de Oficios {anio}'
    ws['A2'].font = Font(bold=True, size=13, color="285973")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f'Fecha de Emisión: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Folio', 'Año', 'Expediente', 'Solicitante', 'Asunto', 'Fecha', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Datos
    for row_idx, o in enumerate(oficios, 6):
        values = [o['folio'], o['anio'], o['expediente'], o['solicitante'],
                  o['asunto'], o['fecha'], o['estado']]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [1, 2, 6, 7]: # Centrar folio, año, fecha y estado
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    user_id = int(get_jwt_identity())
    AuditModel.registrar(user_id, 'EXPORTAR_EXCEL', 'Reportes', None, f"Reporte Excel generado - Año {anio}")
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_oficios_{anio}_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@reportes_bp.route('/auditoria', methods=['GET'])
@admin_requerido
def auditoria():
    """Obtener log de auditoría (solo Admin)."""
    limit = request.args.get('limit', 50, type=int)
    logs = AuditModel.listar(limit)
    return jsonify(logs), 200
